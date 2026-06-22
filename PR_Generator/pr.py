import os

# ====================================================================
# --- AVTOMATİK LİMİT ARTIRICI SİSTEM ---
# ====================================================================
if not os.path.exists(".streamlit"):
    os.makedirs(".streamlit")

with open(".streamlit/config.toml", "w", encoding="utf-8") as f:
    f.write("[server]\nmaxUploadSize = 2000\n")

import streamlit as st
import google.generativeai as genai
import urllib.parse
import requests
import base64
from bs4 import BeautifulSoup
from PIL import Image
import tempfile
import time
import datetime
import json
import pandas as pd
import plotly.express as px

# ====================================================================
# results.json yolu — lokal və Streamlit Cloud üçün
# ====================================================================
def _find_results_file():
    # 1. Repo kökündə (GitHub-dan gəlir)
    repo_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "results.json")
    if os.path.exists(repo_root):
        return repo_root
    # 2. Eyni qovluqda
    local = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results.json")
    if os.path.exists(local):
        return local
    return repo_root  # default (yoxdursa belə göstərir)

AGENT_RESULTS_FILE = _find_results_file()

# ====================================================================
# 1. AI MODELİNİN VƏ SƏHİFƏNİN AYARLANMASI
# ====================================================================
st.set_page_config(
    page_title="PR Monitorinq Dashboard",
    page_icon="📊",
    layout="wide"
)

model = None
try:
    GOOGLE_API_KEY = st.secrets["GOOGLE_API_KEY"]
    genai.configure(api_key=GOOGLE_API_KEY)
    model = genai.GenerativeModel('gemini-1.5-flash')
except Exception:
    pass


# ====================================================================
# 2. KÖMƏKÇİ FUNKSİYALAR
# ====================================================================
@st.cache_data(ttl=86400, show_spinner=False)
def get_image_base64(url):
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code == 200:
            content_type = response.headers.get('Content-Type', 'image/png')
            encoded = base64.b64encode(response.content).decode()
            return f"data:{content_type};base64,{encoded}"
    except Exception:
        pass
    return None


@st.cache_data(ttl=1800, show_spinner=False)
def fetch_article_text(article_url):
    if not article_url:
        return ""
    try:
        resp = requests.get(article_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'html.parser')
            paragraphs = soup.find_all('p')
            return " ".join([p.get_text() for p in paragraphs if len(p.get_text()) > 20])[:3000]
    except Exception:
        pass
    return ""


def _render_cards(df_sub, border_color, bg_color, text_color, show_date=True):
    if df_sub.empty:
        st.info("Bu kateqoriyada xəbər yoxdur.")
        return
    sort_col = "_dt" if "_dt" in df_sub.columns else None
    rows = df_sub.sort_values(sort_col, ascending=False).iterrows() if sort_col else df_sub.iterrows()
    for _, row in rows:
        saved = ""
        if show_date and "_dt" in row and pd.notna(row.get("_dt")):
            saved = row["_dt"].strftime("%d.%m.%Y %H:%M")
        elif show_date and "saved_at" in row and row.get("saved_at"):
            try:
                saved = datetime.datetime.fromisoformat(str(row["saved_at"])).strftime("%d.%m.%Y %H:%M")
            except Exception:
                pass
        kw = row.get("keyword") or row.get("Açar Söz") or ""
        title = row.get("title") or row.get("Mövzu") or ""
        summary = row.get("summary") or row.get("Xülasə (AI)") or ""
        link = row.get("link") or row.get("Link") or "#"
        source = row.get("source") or ""
        st.markdown(
            f'''
            <div style="padding:12px; background-color:{bg_color};
                        border-left:5px solid {border_color};
                        border-radius:4px; margin-bottom:10px;">
                <div style="display:flex; justify-content:space-between;
                            align-items:center; margin-bottom:4px;">
                    <span style="font-size:11px; background-color:{border_color};
                                 color:white; padding:2px 6px; border-radius:3px;">
                        🔑 {kw}</span>
                    <span style="font-size:11px; color:#6b7280;">{saved}</span>
                </div>
                <h5 style="margin:6px 0 4px 0; color:{text_color};">{title}</h5>
                <p style="margin:0 0 6px 0; font-size:13px; color:#4b5563;">
                    <b>AI Xülasə:</b> {summary}</p>
                <div style="display:flex; justify-content:space-between; align-items:center;">
                    <a href="{link}" target="_blank"
                       style="font-size:12px; color:{border_color};
                              font-weight:bold; text-decoration:none;">
                        🔗 Mənbəyə get →</a>
                    <span style="font-size:11px; color:#9ca3af;">{source}</span>
                </div>
            </div>
            ''', unsafe_allow_html=True
        )


# ====================================================================
# BAŞLIQ
# ====================================================================
gerb_url = "https://upload.wikimedia.org/wikipedia/commons/thumb/2/20/Emblem_of_Azerbaijan.svg/500px-Emblem_of_Azerbaijan.svg.png"
hero_logo_src = get_image_base64(gerb_url) or "https://flagcdn.com/w160/az.png"

st.markdown(
    f'''
    <div style="text-align: center; padding: 25px; background-color: #f8f9fa;
                border-radius: 12px; margin-bottom: 25px; border-bottom: 5px solid #1f2937;">
        <img src="{hero_logo_src}" width="75" style="margin-bottom: 12px;">
        <h1 style="color: #1f2937; font-size: 28px; margin-bottom: 6px; font-weight: bold;">
            Operativ PR Monitorinq və Reputasiya Dashboard-u</h1>
        <p style="color: #4b5563; font-size: 16px;">
            Süni İntellekt Dəstəkli Fasiləsiz Böhran Nəzarəti Sistemi</p>
    </div>
    ''',
    unsafe_allow_html=True
)

# ====================================================================
# YAN MENYU
# ====================================================================
st.sidebar.markdown("## 🛠️ İş Rejimi")
rejim = st.sidebar.radio(
    "Zəhmət olmasa bölməni seçin:",
    ["📊 Monitorinq Dashboard", "📝 Press-Reliz Yarat", "📱 Sosial Media Postu Yarat"]
)

if "pr_response" not in st.session_state:
    st.session_state.pr_response = None

COLOR_MAP = {"MÜSBƏT": "#10b981", "NEYTRAL": "#f59e0b", "MƏNFİ": "#ef4444"}

# ====================================================================
# REJİM 1: MONİTORİNQ DASHBOARD
# ====================================================================
if rejim == "📊 Monitorinq Dashboard":
    st.markdown("### 📊 PR Monitorinq Dashboard")

    # ── Bütün məlumatı yüklə ──────────────────────────────────────
    df_raw = pd.DataFrame()
    if os.path.exists(AGENT_RESULTS_FILE):
        try:
            with open(AGENT_RESULTS_FILE, encoding="utf-8") as _f:
                raw = json.load(_f)
            if raw:
                df_raw = pd.DataFrame(raw)
                def _pdt(s):
                    try:
                        return datetime.datetime.fromisoformat(str(s))
                    except Exception:
                        return None
                df_raw["_dt"] = df_raw["saved_at"].apply(_pdt)
        except Exception as e:
            st.warning(f"Agent faylı oxunmadı: {e}")

    if df_raw.empty:
        st.info("📭 Hələ heç bir məlumat yoxdur. Agent.py-ni işə salın.")
        st.stop()

    # ── 🔔 Bildiriş — son 1 saatda yeni mənfi xəbər ───────────────
    one_hour_ago = datetime.datetime.now() - datetime.timedelta(hours=1)
    new_menfi = df_raw[
        (df_raw["sentiment"] == "MƏNFİ") &
        (df_raw["_dt"].notna()) &
        (df_raw["_dt"] >= one_hour_ago)
    ]
    if not new_menfi.empty:
        st.error(
            f"⚠️ **DİQQƏT!** Son 1 saatda **{len(new_menfi)}** yeni mənfi xəbər aşkarlandı! "
            f"Aşağıda 🔴 Mənfi tabına baxın."
        )

    # ── Filtr paneli ──────────────────────────────────────────────
    PERIOD_MAP = {"Son 24 saat": 1, "Son 1 həftə": 7,
                  "Son 1 ay": 30, "Son 1 rüb": 91, "Son 1 il": 365}

    col_p, col_s, col_src, col_q, col_btn = st.columns([2, 2, 2, 3, 1])

    with col_p:
        selected_period = st.selectbox("📅 Dövr:", list(PERIOD_MAP.keys()))
        period_cutoff = datetime.datetime.now() - datetime.timedelta(days=PERIOD_MAP[selected_period])

    with col_s:
        selected_sent = st.selectbox("🎯 Sentiment:", ["Hamısı", "MƏNFİ", "NEYTRAL", "MÜSBƏT"])

    with col_src:
        src_list = ["Hamısı"] + sorted(df_raw["source"].dropna().unique().tolist())
        selected_src = st.selectbox("🌐 Mənbə:", src_list)

    with col_q:
        search_query = st.text_input("🔍 Başlıqda axtar:", placeholder="məs: müavinət...")

    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄", use_container_width=True, help="Yenilə"):
            st.rerun()

    # ── Filtrasiya ────────────────────────────────────────────────
    df_all = df_raw[df_raw["_dt"] >= period_cutoff].copy()
    if selected_sent != "Hamısı":
        df_all = df_all[df_all["sentiment"] == selected_sent]
    if selected_src != "Hamısı":
        df_all = df_all[df_all["source"] == selected_src]
    if search_query.strip():
        q = search_query.strip().lower()
        df_all = df_all[df_all["title"].str.lower().str.contains(q, na=False)]

    if df_all.empty:
        st.info("Bu filtrə uyğun xəbər tapılmadı.")
        st.stop()

    # ── Metriklər ─────────────────────────────────────────────────
    n_menfi   = len(df_all[df_all["sentiment"] == "MƏNFİ"])
    n_neytral = len(df_all[df_all["sentiment"] == "NEYTRAL"])
    n_musbet  = len(df_all[df_all["sentiment"] == "MÜSBƏT"])

    # Əvvəlki dövrlə müqayisə
    prev_start = period_cutoff - datetime.timedelta(days=PERIOD_MAP[selected_period])
    df_prev = df_raw[(df_raw["_dt"] >= prev_start) & (df_raw["_dt"] < period_cutoff)]
    delta_total = len(df_all) - len(df_prev)
    delta_menfi = n_menfi - len(df_prev[df_prev["sentiment"] == "MƏNFİ"])

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("📰 Ümumi xəbər", len(df_all), delta=delta_total)
    m2.metric("🔴 Mənfi", n_menfi, delta=delta_menfi, delta_color="inverse")
    m3.metric("🟡 Neytral", n_neytral)
    m4.metric("🟢 Müsbət", n_musbet)

    st.markdown("---")

    # ── Qrafiklər ─────────────────────────────────────────────────
    col_pie, col_bar = st.columns(2)

    with col_pie:
        st.markdown("#### 🥧 Sentiment bölgüsü")
        counts = df_all["sentiment"].value_counts().reset_index()
        counts.columns = ["Sentiment", "Say"]
        fig_pie = px.pie(counts, values="Say", names="Sentiment",
                         color="Sentiment", color_discrete_map=COLOR_MAP, hole=0.4)
        fig_pie.update_layout(margin=dict(t=10, b=10, l=10, r=10))
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_bar:
        st.markdown("#### 📊 Günlük xəbər axını")
        df_all["_date"] = pd.to_datetime(df_all["_dt"]).dt.date
        daily = df_all.groupby(["_date", "sentiment"]).size().reset_index(name="say")
        fig_bar = px.bar(daily, x="_date", y="say", color="sentiment",
                         color_discrete_map=COLOR_MAP, barmode="stack",
                         labels={"_date": "", "say": "Say", "sentiment": "Sentiment"})
        fig_bar.update_layout(margin=dict(t=10, b=10, l=10, r=10))
        st.plotly_chart(fig_bar, use_container_width=True)

    st.markdown("#### 🔑 Açar söz üzrə bölgü")
    kw_c = df_all.groupby(["keyword", "sentiment"]).size().reset_index(name="say")
    fig_kw = px.bar(kw_c, x="keyword", y="say", color="sentiment",
                    color_discrete_map=COLOR_MAP, barmode="stack",
                    labels={"keyword": "", "say": "Say", "sentiment": "Sentiment"})
    fig_kw.update_layout(xaxis_tickangle=-30, margin=dict(t=10, b=120, l=10, r=10))
    st.plotly_chart(fig_kw, use_container_width=True)

    # ── Xəbər kartları ────────────────────────────────────────────
    st.markdown("---")
    st.markdown(f"#### 📋 {selected_period} — {len(df_all)} xəbər")

    tab_m, tab_n, tab_p = st.tabs([
        f"🔴 Mənfi ({n_menfi})",
        f"🟡 Neytral ({n_neytral})",
        f"🟢 Müsbət ({n_musbet})",
    ])
    with tab_m:
        _render_cards(df_all[df_all["sentiment"] == "MƏNFİ"],
                      "#ef4444", "#fef2f2", "#991b1b")
    with tab_n:
        _render_cards(df_all[df_all["sentiment"] == "NEYTRAL"],
                      "#f59e0b", "#fffbeb", "#92400e")
    with tab_p:
        _render_cards(df_all[df_all["sentiment"] == "MÜSBƏT"],
                      "#10b981", "#f0fdf4", "#065f46")

    # ── Excel / PDF ixrac ─────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📥 Hesabat İxracı")

    import io

    export_df = df_all[["title", "keyword", "sentiment", "source",
                         "summary", "link", "saved_at"]].copy()
    export_df.columns = ["Başlıq", "Açar söz", "Sentiment",
                         "Mənbə", "Xülasə", "Link", "Tarix"]

    col_xl, col_pdf = st.columns(2)

    with col_xl:
        try:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False, sheet_name="Xəbərlər")
                ws = writer.sheets["Xəbərlər"]
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                header_font  = Font(bold=True, color="FFFFFF")
                header_fill  = PatternFill("solid", fgColor="1F4E79")
                sent_colors  = {"MƏNFİ": "FDDCDC", "NEYTRAL": "FFF3CD", "MÜSBƏT": "D4EDDA"}
                for col_idx, col_name in enumerate(export_df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font  = header_font
                    cell.fill  = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for row_idx, row in enumerate(export_df.itertuples(index=False), 2):
                    sent = row[2]
                    fill_color = sent_colors.get(sent)
                    if fill_color:
                        fill = PatternFill("solid", fgColor=fill_color)
                        for col_idx in range(1, len(export_df.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                col_widths = [60, 25, 12, 20, 60, 50, 20]
                for i, w in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w
            st.download_button(
                "📊 Excel (.xlsx) yüklə",
                data=buf.getvalue(),
                file_name=f"pr_hesabat_{selected_period.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except ImportError:
            st.info("Excel üçün: `pip install openpyxl`")

    with col_pdf:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Paragraph, Spacer)
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os as _os

            # DejaVu şrifti var isə qeydiyyat (Unicode dəstəyi üçün)
            _font_dirs = [
                r"C:\Windows\Fonts",
                "/usr/share/fonts/truetype/dejavu",
                "/usr/share/fonts",
            ]
            _font_name = "Helvetica"
            for _d in _font_dirs:
                _path = _os.path.join(_d, "DejaVuSans.ttf")
                if _os.path.exists(_path):
                    try:
                        pdfmetrics.registerFont(TTFont("DejaVuSans", _path))
                        _font_name = "DejaVuSans"
                    except Exception:
                        pass
                    break

            buf_pdf = io.BytesIO()
            doc = SimpleDocTemplate(
                buf_pdf, pagesize=A4,
                leftMargin=1.5*cm, rightMargin=1.5*cm,
                topMargin=2*cm, bottomMargin=2*cm,
            )
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "PDFTitle", parent=styles["Heading1"],
                fontName=_font_name, fontSize=14, spaceAfter=6,
            )
            sub_style = ParagraphStyle(
                "PDFSub", parent=styles["Normal"],
                fontName=_font_name, fontSize=9, textColor=rl_colors.grey,
            )
            cell_style = ParagraphStyle(
                "Cell", parent=styles["Normal"],
                fontName=_font_name, fontSize=7.5, leading=10,
            )

            story = [
                Paragraph("PR Monitorinq Hesabatı", title_style),
                Paragraph(
                    f"{selected_period}  |  Cəmi: {len(export_df)} xəbər  |  "
                    f"Tarix: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}",
                    sub_style,
                ),
                Spacer(1, 0.4*cm),
            ]

            # Cədvəl başlığı
            headers_pdf = ["Başlıq", "Açar söz", "Sentiment", "Mənbə", "Xülasə"]
            col_widths_pdf = [6*cm, 3.5*cm, 2.2*cm, 3*cm, 6.8*cm]
            table_data = [[Paragraph(h, ParagraphStyle(
                "H", fontName=_font_name, fontSize=8,
                textColor=rl_colors.white, fontWeight="bold",
            )) for h in headers_pdf]]

            _sent_rgb = {
                "MƏNFİ":   rl_colors.HexColor("#FDDCDC"),
                "NEYTRAL":  rl_colors.HexColor("#FFF3CD"),
                "MÜSBƏT":   rl_colors.HexColor("#D4EDDA"),
            }
            row_sent_colors = []
            for _, row in export_df.iterrows():
                table_data.append([
                    Paragraph(str(row["Başlıq"])[:120], cell_style),
                    Paragraph(str(row["Açar söz"])[:40], cell_style),
                    Paragraph(str(row["Sentiment"]), cell_style),
                    Paragraph(str(row["Mənbə"])[:30], cell_style),
                    Paragraph(str(row["Xülasə"])[:200], cell_style),
                ])
                row_sent_colors.append(row["Sentiment"])

            tbl = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
            ts = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1F4E79")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), rl_colors.white),
                ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
                ("GRID",       (0, 0), (-1, -1), 0.3, rl_colors.grey),
                ("VALIGN",     (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [rl_colors.white, rl_colors.HexColor("#F7F7F7")]),
            ])
            for r_idx, sent in enumerate(row_sent_colors, 1):
                bg = _sent_rgb.get(sent)
                if bg:
                    ts.add("BACKGROUND", (0, r_idx), (-1, r_idx), bg)
            tbl.setStyle(ts)
            story.append(tbl)

            doc.build(story)
            st.download_button(
                "📄 PDF yüklə",
                data=buf_pdf.getvalue(),
                file_name=f"pr_hesabat_{selected_period.replace(' ','_')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except ImportError:
            st.info("PDF üçün: `pip install reportlab`")

    # ── Alt məlumat ───────────────────────────────────────────────
    try:
        last_ts = max((r.get("saved_at", "") for r in raw), default="")
        if last_ts:
            last_dt = datetime.datetime.fromisoformat(last_ts)
            st.caption(
                f"🕐 Son yeniləmə: {last_dt.strftime('%d.%m.%Y %H:%M')} "
                f"| Arxivdə cəmi: {len(df_raw)} xəbər"
            )
    except Exception:
        pass

# ====================================================================
# REJİM 2: PRESS-RELİZ YARAT
# ====================================================================
elif rejim == "📝 Press-Reliz Yarat":
    with st.form("pr_form"):
        st.markdown('### 🏢 Dövlət / Korporativ Üslubda Press-Reliz')
        sirket_adi = st.text_input("Qurum və ya Şirkət Adı",
                                   placeholder="Məs: Əmək və Əhalinin Sosial Müdafiəsi Nazirliyi...")
        movzu = st.text_area("Xəbərin/Tədbirin Əsas Məzmunu", height=100)
        menbe_url = st.text_input("İstinad üçün Xəbər Linki (İstəyə bağlı)")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.selectbox("Hədəf Kütlə",
                         ["Geniş İctimaiyyət", "Rəsmi Nümayəndələr", "İxtisaslaşmış Media"])
        with col2:
            ton = st.selectbox("Səs Tonu",
                               ["Rəsmi / Dövlət üslubu", "İnformasiya xarakterli", "Analitik"])
        with col3:
            dil = st.selectbox("Dil", ["Azərbaycanca", "English", "Русский"])

        submitted = st.form_submit_button("Sənədi Formalaşdır")

    if submitted and movzu:
        if not model:
            st.error("API açarı tapılmadı!")
        else:
            with st.spinner("⏳ Press-reliz hazırlanır..."):
                uslub = fetch_article_text(menbe_url) if menbe_url else ""
                prompt = (f"Sən {sirket_adi} qurumunun Mətbuat Katibisən. "
                          f"Mövzu: {movzu}. Dil: {dil}. Ton: {ton}. "
                          f"Üslub nümunəsi: {uslub}. Dolğun rəsmi press-reliz sənədi yaz.")
                try:
                    st.session_state.pr_response = model.generate_content(prompt).text
                    st.success("Sənəd uğurla formalaşdırıldı!")
                except Exception as e:
                    st.error(f"Xəta: {e}")

    if st.session_state.pr_response:
        st.markdown("---")
        st.markdown(st.session_state.pr_response)
        st.download_button("📥 Mətni Saxla (.txt)",
                           st.session_state.pr_response, "press_reliz.txt")

# ====================================================================
# REJİM 3: SOSİAL MEDİA POSTU
# ====================================================================
elif rejim == "📱 Sosial Media Postu Yarat":
    st.markdown('### 📱 Sosial Şəbəkə Paylaşımı Hazırlanması')
    sm_text = st.text_area("Xəbərin Mətni",
                           value=st.session_state.pr_response or "", height=150)
    uploaded_file = st.file_uploader("Poster və ya Video Yüklə (2GB-a qədər)",
                                     type=['jpg', 'jpeg', 'png', 'mp4'])

    if st.button("Sosial Media Postunu Hazırla"):
        if not model:
            st.error("API açarı tapılmadı!")
        else:
            with st.spinner("⏳ Post hazırlanır..."):
                prompt = (f"Aşağıdakı məlumata əsasən rəsmi qurumun sosial media hesabı üçün "
                          f"maraqlı, emojili post başlığı (caption) yaz:\n{sm_text}")
                try:
                    if uploaded_file:
                        if uploaded_file.type in ['image/jpeg', 'image/png', 'image/jpg']:
                            response = model.generate_content([prompt, Image.open(uploaded_file)])
                        elif uploaded_file.type == 'video/mp4':
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".mp4") as tmp:
                                tmp.write(uploaded_file.read())
                                tmp_path = tmp.name
                            vid = genai.upload_file(path=tmp_path)
                            while vid.state.name == "PROCESSING":
                                time.sleep(2)
                                vid = genai.get_file(vid.name)
                            response = model.generate_content([prompt, vid])
                            os.unlink(tmp_path)
                    else:
                        response = model.generate_content(prompt)
                    st.markdown("### 📝 Hazırlanan Post:")
                    st.markdown(response.text)
                except Exception as e:
                    st.error(f"Xəta: {e}")

st.markdown(
    "<br><hr><center><p style='color: gray;'>"
    "PR Monitorinq Dashboard | Operativ Analitika Modulu 🤖"
    "</p></center>",
    unsafe_allow_html=True
)
